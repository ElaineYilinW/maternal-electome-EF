"""
Off-paper data exports for Sara's one-off analysis requests.

Each function corresponds to one xlsx artifact the OnnestVsOffnest task
notebook used to produce inline (in a single very long cell). The Sara
section of the notebook collapses to one call per function:

    sara_pup_retrieval(model, ...,         output_xlsx="Onnest_pups.xlsx")
    sara_onnest_loading_inspect(model, ..., output_xlsx="OnNestEF_OnOff.xlsx")
    sara_p3_behavior(model, ...,           output_xlsx="OnNestEF_Behavior_P3.xlsx")

These are not paper figures -- they are convenience dumps Sara requested.
Each writes a multi-sheet workbook and prints a single line on success.

Function provenance (from the OnnestVsOffnest_3band notebook):
    sara_pup_retrieval         <- c53 (~200 lines inline)
    sara_onnest_loading_inspect <- c58 (~150 lines inline)
    sara_p3_behavior           <- c60 (~185 lines inline)

The two pure-exploration cells (c55 "Jan5 score request data prep" and c56
"Jan21 score request apply model") are intentionally NOT ported: they
produced no artifact, only printed counts. Recreate them ad hoc if needed.
"""

import pickle
from typing import Optional, Sequence

import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from data_utils import clean_mouse_id


# =============================================================================
# Helper: load pkl + apply model, return (data_dict, s0_scores)
# =============================================================================

def _load_and_score(model, data_file, x_feature_list=None, x_feature_weights=None):
    """Load a pkl, hstack the configured features, apply the model, return
    (data_dict, s_first_dim)."""
    if x_feature_list is None:
        x_feature_list = ["power", "coh_sq_coherence"]
    if x_feature_weights is None:
        x_feature_weights = [1, 1]

    with open(data_file, "rb") as f:
        data_dict = pickle.load(f)

    X = np.hstack([
        data_dict[feature] * weight
        for feature, weight in zip(x_feature_list, x_feature_weights)
    ])

    model.eval()
    with torch.no_grad():
        _, s_scores = model.predict_proba(X, include_scores=True)

    s_first = s_scores[:, 0] if len(s_scores.shape) > 1 else s_scores
    return data_dict, s_first


def _summary_stats(values):
    """Return per-vector summary stats used by the Sara xlsx tables."""
    n = len(values)
    if n == 0:
        return {'Sample_Count': 0, 'Mean': np.nan, 'Std': np.nan,
                'Min': np.nan, 'Max': np.nan, 'Median': np.nan,
                'Q25': np.nan, 'Q75': np.nan}
    return {
        'Sample_Count': int(n),
        'Mean': float(np.mean(values)),
        'Std': float(np.std(values, ddof=1)) if n > 1 else np.nan,
        'Min': float(np.min(values)),
        'Max': float(np.max(values)),
        'Median': float(np.median(values)),
        'Q25': float(np.percentile(values, 25)),
        'Q75': float(np.percentile(values, 75)),
    }


# =============================================================================
# Sara request 1: pup-retrieval per-mouse loading-score detail
# =============================================================================

def sara_pup_retrieval(model, *,
                       pup_retrieval_data_file,
                       c_mouse_ids, e_mouse_ids,
                       output_xlsx,
                       baseline_n_samples=400,
                       target_period='P4 home'):
    """Per-mouse loading-score time-series at P4 home, split by Sara's
    4-label scheme:

        0 = baseline (first ``baseline_n_samples`` samples per mouse)
        1 = trial, no retrieval        (y_detail == 1)
        2 = retrieval (partial/full)   (y_detail in [3, 4])
        3 = other (post-baseline, non-trial)

    Output workbook has two sheets:
        Individual : side-by-side per-mouse (Loading_Score, Label) columns
        Summary    : per-mouse-per-label N / mean / median / std / sem / etc.

    Defaults match the original notebook (400 samples = 20 min at 3 s/window).
    """
    label_names = {0: 'Baseline', 1: 'Trial (no retrieval)',
                   2: 'Retrieval', 3: 'Other'}

    # ---- Load data + clean ids -----------------------------------------------
    with open(pup_retrieval_data_file, "rb") as f:
        data = pickle.load(f)
    cleaned_mouse_ids = np.array([clean_mouse_id(m) for m in data['mouse_id']])

    c_in = [m for m in c_mouse_ids if m in set(cleaned_mouse_ids)]
    e_in = [m for m in e_mouse_ids if m in set(cleaned_mouse_ids)]
    target_mice = c_in + e_in

    mask = np.isin(cleaned_mouse_ids, target_mice) & (data['period'] == target_period)
    if mask.sum() == 0:
        raise ValueError(f"No samples at period={target_period!r} for the given mice")

    X = np.hstack([data['power'][mask], data['coh_sq_coherence'][mask]])
    y_detail = data['pup_retrieval_detail'][mask]
    mouse_ids = cleaned_mouse_ids[mask]

    # ---- Apply model ---------------------------------------------------------
    model.eval()
    with torch.no_grad():
        _, s_scores = model.predict_proba(X, include_scores=True)
    s0 = s_scores[:, 0] if len(s_scores.shape) > 1 else s_scores

    # ---- Assign 4-class labels per mouse -------------------------------------
    new_labels = np.full(len(y_detail), 3, dtype=int)
    for mid in np.unique(mouse_ids):
        local_idx = np.where(mouse_ids == mid)[0]
        for local_i, global_i in enumerate(local_idx):
            if local_i < baseline_n_samples:
                new_labels[global_i] = 0
            elif y_detail[local_i] == 1:
                new_labels[global_i] = 1
            elif y_detail[local_i] in [3, 4]:
                new_labels[global_i] = 2
            else:
                new_labels[global_i] = 3

    # ---- Write xlsx ----------------------------------------------------------
    mouse_order = sorted(np.unique(mouse_ids))
    wb = Workbook()

    # Sheet 1: Individual (one mouse = 2 columns: score, label)
    ws_ind = wb.active
    ws_ind.title = 'Individual'
    ws_ind['A1'] = (f'Label: 0=Baseline (first {baseline_n_samples} samples per mouse), '
                    f'1=Trial no retrieval, 2=Retrieval (partial/complete), 3=Other')
    for col_offset, mid in enumerate(mouse_order):
        col_score = col_offset * 2 + 1
        col_label = col_offset * 2 + 2
        group = 'C' if mid in c_in else 'E'

        ws_ind.merge_cells(start_row=2, start_column=col_score,
                           end_row=2, end_column=col_label)
        ws_ind.cell(row=2, column=col_score).value = f'{mid} (Grp {group})'
        ws_ind.cell(row=3, column=col_score).value = 'Loading Score'
        ws_ind.cell(row=3, column=col_label).value = 'Label'

        m_mask = mouse_ids == mid
        scores_m = s0[m_mask]
        labels_m = new_labels[m_mask]
        for row_i, (sc, lb) in enumerate(zip(scores_m, labels_m)):
            ws_ind.cell(row=row_i + 4, column=col_score).value = float(sc)
            ws_ind.cell(row=row_i + 4, column=col_label).value = int(lb)

        ws_ind.column_dimensions[get_column_letter(col_score)].width = 16
        ws_ind.column_dimensions[get_column_letter(col_label)].width = 8

    # Sheet 2: Summary (per-mouse per-label stats)
    ws_sum = wb.create_sheet('Summary')
    stat_cols = ['N', 'Mean', 'Median', 'STD', 'SEM', 'Min', 'Max', 'Q25', 'Q75']
    headers = ['Mouse_ID', 'Group', 'Label', 'Label_Name'] + stat_cols
    for col_i, h in enumerate(headers, start=1):
        ws_sum.cell(row=1, column=col_i).value = h

    row_i = 2
    for mid in mouse_order:
        group = 'C' if mid in c_in else 'E'
        scores_m = s0[mouse_ids == mid]
        labels_m = new_labels[mouse_ids == mid]
        for lbl in [0, 1, 2, 3]:
            vals = scores_m[labels_m == lbl]
            n = len(vals)
            std = np.std(vals, ddof=1) if n > 1 else np.nan
            sem = std / np.sqrt(n) if n > 1 else np.nan
            row = [mid, group, lbl, label_names[lbl],
                   n, np.mean(vals) if n else np.nan,
                   np.median(vals) if n else np.nan,
                   std, sem,
                   np.min(vals) if n else np.nan,
                   np.max(vals) if n else np.nan,
                   np.percentile(vals, 25) if n else np.nan,
                   np.percentile(vals, 75) if n else np.nan]
            for col_i, v in enumerate(row, start=1):
                ws_sum.cell(row=row_i, column=col_i).value = (
                    None if (isinstance(v, float) and np.isnan(v)) else v
                )
            row_i += 1

    for i, w in enumerate([15, 8, 8, 22, 6, 12, 12, 12, 12, 12, 12, 12, 12], start=1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_xlsx)
    print(f"  Sara pup retrieval written: {output_xlsx}  "
          f"({len(mouse_order)} mice, Individual + Summary sheets)")


# =============================================================================
# Sara request 2: onnest loading inspect (per-period per-mouse per-onnest)
# =============================================================================

def sara_onnest_loading_inspect(model, *, training_data_file, output_xlsx):
    """Apply model to every (period, mouse, onnest_label) cell of the
    8-region pkl and write the 4-sheet xlsx (Full_Statistics, Mean, Median,
    Raw_Scores).
    """
    data, s0 = _load_and_score(model, training_data_file)

    periods = np.unique(data['period'])
    mice = np.unique(data['mouse_id'])
    onnest_vals = np.unique(data['onnest_label'])

    # ---- Full statistics row per (period, mouse, onnest) ---------------------
    rows = []
    for p in periods:
        for m in mice:
            for o in onnest_vals:
                cell_mask = (data['period'] == p) & (data['mouse_id'] == m) & (data['onnest_label'] == o)
                vals = s0[cell_mask]
                stats = _summary_stats(vals)
                rows.append({'Period': p, 'Mouse_ID': m, 'Onnest_Label': o, **stats})
    df = pd.DataFrame(rows)

    # ---- Raw scores (long, then pivot wide) ----------------------------------
    df_raw = pd.DataFrame({
        'Period': data['period'],
        'Mouse_ID': data['mouse_id'],
        'Onnest_Label': data['onnest_label'],
        'Score': s0,
    })
    df_raw['Sample_Index'] = df_raw.groupby(['Period', 'Mouse_ID', 'Onnest_Label']).cumcount()
    df_raw_pivot = df_raw.pivot_table(
        index=['Period', 'Onnest_Label', 'Sample_Index'],
        columns='Mouse_ID', values='Score', aggfunc='first',
    ).reset_index()

    # ---- Mean / Median pivots ------------------------------------------------
    df_mean_pivot = df.pivot_table(
        index=['Period', 'Onnest_Label'], columns='Mouse_ID', values='Mean',
    ).reset_index()
    df_median_pivot = df.pivot_table(
        index=['Period', 'Onnest_Label'], columns='Mouse_ID', values='Median',
    ).reset_index()

    # ---- Write 4-sheet workbook ----------------------------------------------
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Full_Statistics', index=False)
        df_mean_pivot.to_excel(writer, sheet_name='Mean', index=False)
        df_median_pivot.to_excel(writer, sheet_name='Median', index=False)
        df_raw_pivot.to_excel(writer, sheet_name='Raw_Scores', index=False)

    print(f"  Sara onnest loading inspect written: {output_xlsx}  "
          f"({len(df)} (period, mouse, onnest) cells, {len(df_raw)} raw samples)")


# =============================================================================
# Sara request 3: P3 behavior summary (licking / nursing)
# =============================================================================

def sara_p3_behavior(model, *, training_data_file, output_xlsx):
    """Apply model to the P3 lick/selfgroom/nurse pkl and summarize per-mouse
    score statistics restricted to ``licking == 1`` and ``nursing == 1``.

    The pkl uses ``-1`` as the "not applicable" sentinel in the nursing field;
    we convert those to NaN before filtering.
    """
    data, s0 = _load_and_score(model, training_data_file)

    # nursing: -1 -> NaN
    nursing_labels = np.asarray(data['nursing'], dtype=float)
    nursing_labels[nursing_labels == -1] = np.nan

    df_raw = pd.DataFrame({
        'Mouse_ID': data['mouse_id'],
        'Score': s0,
        'Onnest': data['onnest_raw'],
        'Licking': data['licking'],
        'Nursing': nursing_labels,
        'Selfgroom': data['selfgroom'],
    })

    # Per-mouse summary restricted to a behavior == 1
    def _per_mouse_summary(filter_mask):
        sub = df_raw[filter_mask]
        rows = []
        for m in sorted(df_raw['Mouse_ID'].unique()):
            vals = sub.loc[sub['Mouse_ID'] == m, 'Score'].values
            rows.append({'Mouse_ID': m, **_summary_stats(vals)})
        return pd.DataFrame(rows)

    df_licking_summary = _per_mouse_summary(df_raw['Licking'] == 1)
    df_nursing_summary = _per_mouse_summary(df_raw['Nursing'] == 1)

    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df_raw.to_excel(writer, sheet_name='Raw_Data', index=False)
        df_licking_summary.to_excel(writer, sheet_name='Summary_Licking', index=False)
        df_nursing_summary.to_excel(writer, sheet_name='Summary_Nursing', index=False)

    n_lick = (df_raw['Licking'] == 1).sum()
    n_nurse = (df_raw['Nursing'] == 1).sum()
    print(f"  Sara P3 behavior written: {output_xlsx}  "
          f"(licking=1: {n_lick}, nursing=1: {n_nurse} samples)")
