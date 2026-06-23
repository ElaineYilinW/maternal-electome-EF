"""
Dataset assembly: feature pkls + behavior labels -> analysis-ready combined pkls.

This module is the DOWNSTREAM half of the preprocessing pipeline. It assumes
per-mouse spectral feature pkls already exist (produced upstream by
``src/lfp_features.py``) and is responsible for:

  - generating per-window behavior labels from xlsx files
  - adding those labels to per-mouse pkls in place
  - aggregating per-mouse pkls into a single combined pkl
  - filtering to a target mouse cohort and trimming each recording
  - the Sara Jan 6 alignment (merging the nursing field across two datasets)

No project-specific constants (window length, target mouse list, etc.) are
defined at module level. Functions that need them accept them as parameters;
the notebook is the single source of truth for configuration values.

Layout:
    Layer 0 -- utility helpers (path / id / xlsx parsing)
    Layer 1 -- window label generation (binary onnest, majority overlap, 4-level pup)
    Layer 2 -- adding labels to per-mouse pkls (writes pkls in place)
    Layer 3 -- aggregation, filter + trim
    Layer 4 -- Sara nursing alignment
"""

import os
import re
import pickle
import warnings

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ====================================================================
# Layer 0: Utility helpers
# ====================================================================

def get_filenames(label_dir, label_suffix):
    """Return sorted list of files in ``label_dir`` ending with ``label_suffix``.

    Skips Office temp files starting with ``~$``. Prints a count summary.
    """
    assert os.path.exists(label_dir), f"{label_dir} doesn't exist!"
    fns = [
        os.path.join(label_dir, fn)
        for fn in sorted(os.listdir(label_dir))
        if fn.endswith(label_suffix) and not fn.startswith("~$")
    ]
    if len(fns) == 0:
        warnings.warn(f"No files in {label_dir}!")
    else:
        print(f"{len(fns)} files in {label_dir}")
    return fns


def canonical_id(raw_id):
    """Convert any mouse_id variant to canonical short form ``CX_ELSYY``.

    Examples::

        MouseC1F3ELS32 -> C1_ELS32
        MouseC1ELS32   -> C1_ELS32
        C1F3ELS32      -> C1_ELS32
        C1ELS32        -> C1_ELS32
    """
    patterns = [
        r'Mouse([CE]\d+)(?:F\d+)?(ELS\d+)',
        r'([CE]\d+)(?:F\d+)?(ELS\d+)',
    ]
    for p in patterns:
        m = re.match(p, str(raw_id))
        if m:
            return f"{m.group(1)}_{m.group(2)}"
    return raw_id


def strip_mouse_prefix(raw_id):
    """Remove ``Mouse`` prefix only (keep F-generation info).

    ``MouseC1F3ELS32`` -> ``C1F3ELS32``.

    Used by trim helpers when the combined pkl stores full ``Mouse...`` form
    but the target mice list is the prefix-stripped form.
    """
    return str(raw_id).replace('Mouse', '')


def extract_key_from_pkl(filename):
    """Extract canonical mouse id from a per-mouse pkl filename.

    Returns the canonical short form (``CX_ELSYY``) or ``None`` on no match.
    """
    basename = os.path.basename(filename)
    m = re.search(
        r'Mouse([CE]\d+)(?:F\d+)?(ELS\d+)_(?:P\d+(?:\s+home)?|Ges|Pre\s+\w+)\.pkl',
        basename)
    if m:
        return f"{m.group(1)}_{m.group(2)}"
    m = re.search(r'Mouse([CE]\d+)(?:F\d+)?(ELS\d+)', basename)
    if m:
        return f"{m.group(1)}_{m.group(2)}"
    return None


def extract_key_from_behavior(filename):
    """Extract canonical mouse id from a behavior xlsx filename.

    Format: ``C1_ELS32.xlsx`` -> ``C1_ELS32``. Returns ``None`` on no match.
    """
    basename = os.path.basename(filename)
    m = re.match(r'([CE]\d+_ELS\d+)\.xlsx', basename)
    if m:
        return m.group(1)
    return None


def read_behavior_file(file_path):
    """Read a behavior xlsx with ``START``/``STOP`` columns.

    Returns a DataFrame with float ``start`` and ``stop`` columns (seconds).
    Returns an empty DataFrame if ``START`` not found, or ``None`` if missing.
    """
    if not os.path.exists(file_path):
        return None
    df = pd.read_excel(file_path, header=None)
    start_row = start_col = stop_col = None
    for i, row in df.iterrows():
        for j, cell in enumerate(row):
            if isinstance(cell, str) and cell.upper() == 'START':
                start_row, start_col = i, j
                stop_col = j + 1
                break
        if start_row is not None:
            break
    if start_row is None:
        return pd.DataFrame({'start': [], 'stop': []})
    starts = df.iloc[start_row+1:, start_col].dropna().astype(float)
    stops  = df.iloc[start_row+1:, stop_col ].dropna().astype(float)
    return pd.DataFrame({'start': starts.values, 'stop': stops.values})


def read_behavior_file_with_color(file_path):
    """Read P4 pup retrieval xlsx with font color information.

    Red font (RGB ``FFFF0000`` or ``FF0000``) -> partial retrieval (label=3).
    Any other color (black, theme, or unset) -> successful retrieval (label=4).

    Returns DataFrame with columns ``start``, ``stop``, ``label``.
    """
    if not os.path.exists(file_path):
        return None
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active
    start_row = start_col = stop_col = None
    for i, row in enumerate(ws.iter_rows(values_only=False), start=1):
        for j, cell in enumerate(row, start=1):
            if cell.value and isinstance(cell.value, str) and cell.value.upper() == 'START':
                start_row, start_col, stop_col = i, j, j + 1
                break
        if start_row is not None:
            break
    if start_row is None:
        wb.close()
        return pd.DataFrame({'start': [], 'stop': [], 'label': []})

    behaviors = []
    for row_idx in range(start_row + 1, ws.max_row + 1):
        start_cell = ws.cell(row=row_idx, column=start_col)
        stop_cell  = ws.cell(row=row_idx, column=stop_col)
        if start_cell.value is None or stop_cell.value is None:
            continue
        try:
            start_time = float(start_cell.value)
            stop_time  = float(stop_cell.value)
            is_partial = False
            color = start_cell.font.color if start_cell.font else None
            if color and hasattr(color, 'type') and color.type == 'rgb':
                if hasattr(color, 'rgb') and isinstance(color.rgb, str):
                    if color.rgb.upper() in ('FFFF0000', 'FF0000'):
                        is_partial = True
            behaviors.append({
                'start': start_time, 'stop': stop_time,
                'label': 3 if is_partial else 4,
            })
        except (ValueError, TypeError):
            continue
    wb.close()
    return pd.DataFrame(behaviors)


def read_trial_times_excel(excel_path):
    """Read the shared P4 pup retrieval trial-times spreadsheet.

    Each sheet name matches a mouse (e.g. ``C6_ELS42``). Trial times live in
    columns G-I (index 6-8), rows 1-11 with row 1 as header.

    Returns dict ``{mouse_key: DataFrame with 'Trial','Start','End'}``.
    """
    trial_dict = {}
    try:
        xls = pd.ExcelFile(excel_path)
        for sheet_name in xls.sheet_names:
            m = re.match(r'([CE]\d+_ELS\d+)', sheet_name)
            if not m:
                continue
            mouse_key = m.group(1)
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            if df.shape[0] >= 11 and df.shape[1] >= 9:
                trial_data = df.iloc[0:11, 6:9].copy()
                trial_data.columns = ['Trial', 'Start', 'End']
                trial_data = trial_data.iloc[1:].copy()
                trial_data['Start'] = pd.to_numeric(trial_data['Start'], errors='coerce')
                trial_data['End']   = pd.to_numeric(trial_data['End'],   errors='coerce')
                trial_data = trial_data.dropna(subset=['Start', 'End'])
                trial_dict[mouse_key] = trial_data
    except Exception as e:
        print(f"Error reading trial times file: {e}")
    return trial_dict


# ====================================================================
# Layer 1: Window label generation
# ====================================================================

def generate_onnest_labels_binary(n_window, window_duration, onnest_data):
    """Convert on-nest bouts to binary window labels.

    A window is labeled 1 if **>= half** of its duration is covered by
    on-nest bouts; 0 otherwise.

    Returns dict ``{'onnest_label': np.ndarray of shape (n_window,)}``.
    """
    labels = np.zeros(n_window, dtype=int)
    current_idx = 0
    start_col = stop_col = None
    for col in onnest_data.columns:
        cu = str(col).upper()
        if cu == 'START':
            start_col = col
        elif cu == 'STOP':
            stop_col = col
    if start_col is None or stop_col is None:
        raise ValueError("Cannot find START or STOP column")

    for i in range(n_window):
        ws = i * window_duration
        we = (i + 1) * window_duration
        in_nest = 0.0
        while current_idx < len(onnest_data):
            s = onnest_data.iloc[current_idx][start_col]
            t = onnest_data.iloc[current_idx][stop_col]
            if t <= ws:
                current_idx += 1
                continue
            if s >= we:
                break
            os_ = max(ws, s)
            oe = min(we, t)
            if os_ < oe:
                in_nest += (oe - os_)
            if t <= we:
                current_idx += 1
            else:
                break
        if in_nest >= window_duration / 2:
            labels[i] = 1
    return {'onnest_label': labels}


def behavior_to_window_labels(behavior_df, num_windows, window_duration):
    """Convert behavior intervals to binary window labels using majority overlap.

    A window is labeled 1 if the total overlap with any behavior bout is
    **> half** of the window duration. (Note the strict-greater-than: this is
    the historical convention used for the ``onnest_raw`` / ``licking`` /
    ``selfgroom`` / ``nursing`` labels, which differs from
    ``generate_onnest_labels_binary`` by one boundary case.)

    Returns np.ndarray of shape ``(num_windows,)`` with 0/1 values.
    """
    if behavior_df is None or behavior_df.empty:
        return np.zeros(num_windows, dtype=int)
    labels = np.zeros(num_windows, dtype=int)
    for _, row in behavior_df.iterrows():
        s = row['start']
        t = row['stop']
        if s >= num_windows * window_duration:
            continue
        sw = int(s / window_duration)
        ew = min(int(t / window_duration) + 1, num_windows)
        for w in range(sw, ew):
            ws = w * window_duration
            we = (w + 1) * window_duration
            ol = max(0.0, min(t, we) - max(s, ws))
            if ol > window_duration / 2:
                labels[w] = 1
    return labels


def behavior_to_window_labels_detailed(behavior_df, trial_df, num_windows, window_duration):
    """Generate 4-level pup retrieval labels per window.

    Labels:

      0 -- no trial (default, window outside any trial)
      1 -- trial, no retrieval (in a trial but no retrieval bout)
      3 -- partial retrieval (red font in xlsx)
      4 -- successful retrieval (black font in xlsx)

    Priority when bouts overlap a window: 4 > 3 > 1 > 0.

    Returns np.ndarray of shape ``(num_windows,)`` with values in ``{0,1,3,4}``.
    """
    labels = np.zeros(num_windows, dtype=int)
    # Step 1: mark trial windows as 1
    if trial_df is not None and not trial_df.empty:
        for _, row in trial_df.iterrows():
            s = row['Start']
            e = row['End']
            if pd.isna(s) or pd.isna(e) or s >= num_windows * window_duration:
                continue
            sw = int(s / window_duration)
            ew = min(int(e / window_duration) + 1, num_windows)
            for w in range(sw, ew):
                ws = w * window_duration
                we = (w + 1) * window_duration
                if max(0.0, min(e, we) - max(s, ws)) > 0:
                    labels[w] = 1
    # Step 2: overlay retrieval bouts (priority 4 > 3 > 1)
    if behavior_df is not None and not behavior_df.empty:
        for _, row in behavior_df.iterrows():
            s = row['start']
            t = row['stop']
            lbl = row['label']
            if s >= num_windows * window_duration:
                continue
            sw = int(s / window_duration)
            ew = min(int(t / window_duration) + 1, num_windows)
            for w in range(sw, ew):
                ws = w * window_duration
                we = (w + 1) * window_duration
                if max(0.0, min(t, we) - max(s, ws)) > 0:
                    if labels[w] < lbl:
                        labels[w] = lbl
    return labels


# ====================================================================
# Layer 2: Add labels to per-mouse pkls (in-place writes)
# ====================================================================

def add_onnest_labels_to_pkls(pkl_files, onnest_files, window_duration,
                               label_key='onnest_label'):
    """Add binary on-nest labels in-place to per-mouse P1/P3/P8/P14 pkls.

    For each pkl, finds matching onnest xlsx by mouse id (canonical short form)
    and writes a binary label vector (>= half-window overlap) to
    ``pkl[label_key]``. The pkl is overwritten in place.

    Args:
        pkl_files: list of per-mouse pkl paths
        onnest_files: list of per-mouse onnest xlsx paths
        window_duration: window length in seconds (must match the value used
            when the spectral features were computed)
        label_key: pkl field to write into (default ``'onnest_label'``)

    Returns: list of pkl paths that were successfully updated.
    """
    onnest_dict = {}
    for fn in onnest_files:
        k = extract_key_from_behavior(fn)
        if k:
            onnest_dict[k] = fn

    matched = []
    for pkl_file in pkl_files:
        basename = os.path.basename(pkl_file)
        m = re.match(r'Mouse([CE]\d+)(?:F\d+)?(ELS\d+)_P\d+\.pkl', basename)
        if not m:
            continue
        key = f"{m.group(1)}_{m.group(2)}"
        if key not in onnest_dict:
            continue
        try:
            with open(pkl_file, 'rb') as f:
                data = pickle.load(f)
            onnest_data = pd.read_excel(onnest_dict[key])
            n_window = data['power'].shape[0]
            result = generate_onnest_labels_binary(n_window, window_duration, onnest_data)
            data[label_key] = result['onnest_label']
            with open(pkl_file, 'wb') as f:
                pickle.dump(data, f)
            matched.append(pkl_file)
        except Exception as e:
            print(f"  Error processing {basename}: {e}")
    print(f"Updated {len(matched)} pkls with '{label_key}'")
    return matched


def add_behavior_labels_to_pkls(pkl_files, behavior_files_dict, window_duration):
    """Add binary behavior labels (majority overlap) in-place to per-mouse pkls.

    A pkl is processed if at least one behavior xlsx matches by canonical
    mouse id. For each match, the resulting binary label vector is stored
    as ``pkl[label_name]``.

    Args:
        pkl_files: list of per-mouse P3 pkl paths
        behavior_files_dict: dict mapping ``label_name -> list of xlsx paths``,
            e.g. ``{'onnest_raw': onnest_fns, 'licking': lick_fns, ...}``
        window_duration: window length in seconds

    Returns: list of pkl paths that were successfully updated.
    """
    label_to_dict = {}
    for label_name, files in behavior_files_dict.items():
        d = {}
        for f in files:
            k = extract_key_from_behavior(f)
            if k:
                d[k] = f
        label_to_dict[label_name] = d
        print(f"  Loaded {len(d)} files for '{label_name}'")

    updated = []
    for pkl_file in pkl_files:
        pkl_key = extract_key_from_pkl(pkl_file)
        if not pkl_key:
            continue
        if not any(pkl_key in d for d in label_to_dict.values()):
            continue
        try:
            with open(pkl_file, 'rb') as f:
                data = pickle.load(f)
        except Exception as e:
            print(f"  Cannot load {os.path.basename(pkl_file)}: {e}")
            continue
        if 'power' not in data:
            continue
        n_window = data['power'].shape[0]
        added = []
        for label_name, lookup in label_to_dict.items():
            if pkl_key in lookup:
                bdf = read_behavior_file(lookup[pkl_key])
                labels = behavior_to_window_labels(bdf, n_window, window_duration)
                data[label_name] = labels
                added.append(label_name)
        if added:
            with open(pkl_file, 'wb') as f:
                pickle.dump(data, f)
            updated.append(pkl_file)
    print(f"Updated {len(updated)} pkls with behavior labels")
    return updated


def add_pup_retrieval_detail_to_pkls(pkl_files, pup_files, trial_excel_path,
                                      window_duration):
    """Add 4-level ``pup_retrieval_detail`` field in-place to per-mouse P4 home pkls.

    Pkls without a matching pup xlsx and without a trial row for the mouse are
    skipped silently.

    Args:
        pkl_files: list of P4 home pkl paths
        pup_files: list of per-mouse pup retrieval xlsx paths
        trial_excel_path: path to the shared P4 trial-times xlsx
        window_duration: window length in seconds

    Returns: list of pkl paths that were successfully updated.
    """
    pup_dict = {}
    for f in pup_files:
        k = extract_key_from_behavior(f)
        if k:
            pup_dict[k] = f
    trial_dict = read_trial_times_excel(trial_excel_path)
    print(f"  Loaded {len(pup_dict)} pup files, {len(trial_dict)} trial-time sheets")

    updated = []
    for pkl_file in pkl_files:
        pkl_key = extract_key_from_pkl(pkl_file)
        if not pkl_key:
            continue
        pup_file = pup_dict.get(pkl_key)
        trial_df = trial_dict.get(pkl_key)
        has_pup   = pup_file is not None
        has_trial = trial_df is not None and not trial_df.empty
        if not has_pup and not has_trial:
            continue
        try:
            with open(pkl_file, 'rb') as f:
                data = pickle.load(f)
        except Exception as e:
            print(f"  Cannot load {os.path.basename(pkl_file)}: {e}")
            continue
        if 'power' not in data:
            continue
        n_window = data['power'].shape[0]
        bdf = read_behavior_file_with_color(pup_file) if has_pup else None
        tdf = trial_df if has_trial else None
        labels = behavior_to_window_labels_detailed(bdf, tdf, n_window, window_duration)
        data['pup_retrieval_detail'] = labels
        with open(pkl_file, 'wb') as f:
            pickle.dump(data, f)
        updated.append(pkl_file)
    print(f"Updated {len(updated)} P4 pkls with 'pup_retrieval_detail'")
    return updated


# ====================================================================
# Layer 3: Aggregation / filter / trim
# ====================================================================

def aggregate_per_mouse_pkls(pkl_files, output_path, label_keys=(), verbose=True):
    """Concatenate per-mouse pkls into a single combined pkl.

    Always concatenated: ``power``, ``coh_sq_coherence``, ``mouse_id``, ``period``.
    Additional fields listed in ``label_keys`` are also concatenated.
    Static metadata (``freq_band``, ``region``, ``region_pair``) is taken from
    the first pkl encountered.

    Pkls missing any of the required ``label_keys`` are SKIPPED with a warning.

    Returns: the combined dict (also pickled to ``output_path``).
    """
    combined = {
        'power': [], 'coh_sq_coherence': [],
        'mouse_id': [], 'period': [],
        'freq_band': None, 'region': None, 'region_pair': None,
    }
    for k in label_keys:
        combined[k] = []

    n_added = 0
    skipped = []
    for i, fp in enumerate(pkl_files):
        try:
            with open(fp, 'rb') as f:
                data = pickle.load(f)
        except Exception as e:
            skipped.append((fp, str(e)))
            continue
        missing = [k for k in label_keys if k not in data]
        if missing:
            skipped.append((fp, f"missing keys: {missing}"))
            continue
        combined['power'].append(data['power'])
        combined['coh_sq_coherence'].append(data['coh_sq_coherence'])
        combined['mouse_id'].append(data['mouse_id'])
        combined['period'].append(data['period'])
        for k in label_keys:
            combined[k].append(data[k])
        if combined['freq_band'] is None:
            combined['freq_band']  = data['freq_band']
            combined['region']     = data['region']
            combined['region_pair'] = data['region_pair']
        n_added += 1
        if verbose and (i + 1) % 20 == 0:
            print(f"  Processed {i+1}/{len(pkl_files)} files")

    array_keys = ['power', 'coh_sq_coherence', 'mouse_id', 'period'] + list(label_keys)
    for k in array_keys:
        combined[k] = np.concatenate(combined[k], axis=0)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'wb') as f:
        pickle.dump(combined, f)

    print(f"Aggregated {n_added}/{len(pkl_files)} pkls -> {output_path}")
    print(f"  power={combined['power'].shape}, coh={combined['coh_sq_coherence'].shape}")
    for k in label_keys:
        print(f"  {k}: shape={combined[k].shape}")
    if skipped:
        print(f"  Skipped {len(skipped)} files:")
        for fp, reason in skipped[:5]:
            print(f"    - {os.path.basename(fp)}: {reason}")
        if len(skipped) > 5:
            print(f"    ... and {len(skipped) - 5} more")
    return combined


def filter_and_trim_data(input_path, output_path, mice_to_keep, trim_criteria,
                         id_transform=None):
    """Filter a combined pkl to a target mouse set, then trim each recording.

    Each ``(mouse_id, period)`` recording is trimmed to the first ``N`` samples
    according to ``trim_criteria``. Periods not in ``trim_criteria`` are kept
    untrimmed (all samples retained).

    Args:
        input_path: path to the combined pkl
        output_path: path to save the filtered+trimmed pkl
        mice_to_keep: list of mouse IDs (must match the output of ``id_transform``,
            or the raw ``mouse_id`` strings if ``id_transform`` is None)
        trim_criteria: dict ``{period: max_samples_per_recording or None}``,
            e.g. ``{'P1': 4800, 'P3': 3600, 'P8': 2400}``
        id_transform: optional callable applied to each raw ``mouse_id`` before
            matching against ``mice_to_keep``. Common choices: ``None`` (raw),
            ``canonical_id`` (full -> short), ``strip_mouse_prefix`` (remove
            ``Mouse`` prefix only).

    Returns: the trimmed dict (also pickled to ``output_path``).
    """
    with open(input_path, 'rb') as f:
        full = pickle.load(f)

    raw_ids = np.array(full['mouse_id'])
    if id_transform is not None:
        match_ids = np.array([id_transform(m) for m in raw_ids])
    else:
        match_ids = raw_ids

    df = pd.DataFrame({
        'mouse_id': match_ids,
        'period':   full['period'],
        'index':    np.arange(len(raw_ids)),
    })
    df = df[df['mouse_id'].isin(mice_to_keep)].copy()
    print(f"  After mouse filter: {len(df)} samples, {df['mouse_id'].nunique()} mice")

    keep_indices = []
    for (mid, p), group in df.groupby(['mouse_id', 'period']):
        max_samples = trim_criteria.get(p)
        if max_samples is not None:
            keep_indices.extend(group['index'].values[:max_samples])
        else:
            keep_indices.extend(group['index'].values)
    keep_indices = np.array(sorted(keep_indices), dtype=np.int64)
    print(f"  After trim: {len(keep_indices)} samples")

    trimmed = {}
    n_total = len(raw_ids)
    for k, v in full.items():
        if k in ('freq_band', 'region', 'region_pair'):
            trimmed[k] = v
        elif hasattr(v, 'shape') and len(v) == n_total:
            if k == 'mouse_id' and id_transform is not None:
                trimmed[k] = match_ids[keep_indices]
            else:
                trimmed[k] = v[keep_indices]
        else:
            trimmed[k] = v

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'wb') as f:
        pickle.dump(trimmed, f)

    print(f"Saved trimmed dataset to {output_path}")
    return trimmed


# ====================================================================
# Layer 4: Special operations
# ====================================================================

def align_nursing_complete(with_nursing_path, no_nursing_path, output_path,
                           target_mice, max_samples=3600):
    """Sara Jan 6 request: merge nursing field across two datasets and trim.

    Builds a "complete" dataset combining:

      - all 14 mice (from ``no_nursing_path``)
      - the ``nursing`` field, populated only for mice present in
        ``with_nursing_path``; mice without nursing data get ``nursing = -1``
        (NA marker)

    Then trims each ``(mouse, period)`` recording to the first ``max_samples``.

    Args:
        with_nursing_path: pkl with ``nursing`` field, smaller mouse subset
        no_nursing_path: pkl with all 14 mice, no ``nursing`` field
        output_path: where to save the aligned + trimmed result
        target_mice: list of mouse IDs to keep (in the form used in the input
            pkls, typically the full ``Mouse...`` strings)
        max_samples: max samples per ``(mouse, period)`` recording

    Returns: the final dict (also pickled).
    """
    with open(with_nursing_path, 'rb') as f:
        with_nursing = pickle.load(f)
    with open(no_nursing_path, 'rb') as f:
        no_nursing = pickle.load(f)

    selected = set(target_mice)
    mask = np.isin(no_nursing['mouse_id'], list(selected))
    print(f"  Filter to target mice: {len(no_nursing['mouse_id'])} -> {mask.sum()} samples")

    filtered = {}
    n_total = len(no_nursing['mouse_id'])
    for k, v in no_nursing.items():
        if isinstance(v, np.ndarray) and len(v) == n_total:
            filtered[k] = v[mask]
        else:
            filtered[k] = v

    nursing_array = np.full(len(filtered['mouse_id']), -1, dtype=np.int64)
    mice_with = set(np.unique(with_nursing['mouse_id']))

    matched_recordings = 0
    mismatched = []
    recordings = sorted(set(zip(filtered['period'], filtered['mouse_id'])))
    for period, mouse in recordings:
        if mouse not in mice_with:
            continue
        idx_f = np.where((filtered['period'] == period) &
                         (filtered['mouse_id'] == mouse))[0]
        idx_w = np.where((with_nursing['period'] == period) &
                         (with_nursing['mouse_id'] == mouse))[0]
        if len(idx_w) == 0:
            continue
        if len(idx_f) == len(idx_w):
            nursing_array[idx_f] = with_nursing['nursing'][idx_w]
            matched_recordings += 1
        else:
            mismatched.append((mouse, period, len(idx_w), len(idx_f)))
    filtered['nursing'] = nursing_array
    print(f"  Aligned nursing for {matched_recordings} recordings")
    if mismatched:
        print(f"  WARNING: {len(mismatched)} recordings had sample-count mismatch")

    # Trim per recording
    keep = []
    for period, mouse in recordings:
        m = (filtered['period'] == period) & (filtered['mouse_id'] == mouse)
        idx = np.where(m)[0]
        keep.extend(idx[:max_samples])
    keep = np.array(sorted(keep))

    final = {}
    n_filtered = len(filtered['mouse_id'])
    for k, v in filtered.items():
        if isinstance(v, np.ndarray) and len(v) == n_filtered:
            final[k] = v[keep]
        else:
            final[k] = v

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'wb') as f:
        pickle.dump(final, f)

    counts = {
        0:  int(np.sum(final['nursing'] == 0)),
        1:  int(np.sum(final['nursing'] == 1)),
        -1: int(np.sum(final['nursing'] == -1)),
    }
    print(f"  Saved {len(final['mouse_id'])} samples to {output_path}")
    print(f"  Nursing distribution: 0={counts[0]}, 1={counts[1]}, -1 (NA)={counts[-1]}")
    return final


# ====================================================================
# Module-level sanity check
# ====================================================================

if __name__ == '__main__':
    # Pure-function checks; no file I/O.
    assert canonical_id('MouseC1F3ELS32') == 'C1_ELS32'
    assert canonical_id('MouseC1ELS32')   == 'C1_ELS32'
    assert canonical_id('C1F3ELS32')      == 'C1_ELS32'
    assert canonical_id('C1ELS32')        == 'C1_ELS32'
    assert strip_mouse_prefix('MouseC1F3ELS32') == 'C1F3ELS32'
    assert extract_key_from_behavior('/foo/C1_ELS32.xlsx') == 'C1_ELS32'
    assert extract_key_from_pkl('/foo/MouseC1F3ELS32_P3.pkl') == 'C1_ELS32'
    assert extract_key_from_pkl('/foo/MouseC1F3ELS32_P4 home.pkl') == 'C1_ELS32'
    assert extract_key_from_pkl('/foo/MouseE2ELS3_P1.pkl') == 'E2_ELS3'

    # Synthetic label test: binary onnest with window_duration=3.0
    df_bouts = pd.DataFrame({'START': [0.0, 10.0], 'STOP': [4.0, 18.0]})
    r = generate_onnest_labels_binary(8, 3.0, df_bouts)
    assert r['onnest_label'][0] == 1
    assert r['onnest_label'][1] == 0
    assert r['onnest_label'][3] == 1
    assert r['onnest_label'][4] == 1
    assert r['onnest_label'][5] == 1

    print("dataset_assembly.py sanity checks passed.")
